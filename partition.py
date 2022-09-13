import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import messagebox

#######################################변수 및 동작 설명##############################################
# wb    에 Write_List 엑셀 호출 / ws        에 sheet 호출
# key_i 에 School     엑셀 호출 / key_sheet 에 sheet 호출
# row_list / 반복문을 통해서 Write_List에 기록된 Cell값을 불러옴 
# modify_list / row_list 자체는 비교할 때 쓰이다보니 modify_list로 복제 후에 가공해서 데이터를 입력 / append를 이용해서 값 추가
# school_list / Write_List에 적힌 학교 추출 (오름차순)
# error_school_list / excel 불러올 때 어디서 에러가 났는지 확인을 위한 변수
# modify_school_list / school_list 자체는 비교할 때 쓰이다보니 modify_school_list로 복제 후에 가공해서 데이터를 입력 / 3으로 잘라서 초-등학교, 중-학교로 붙이게 설계
# sheet_list  / Write_List에 적힌 날짜 추출 (오름차순) / ex)202207
# sheet_list_name / 생성할 Sheet명 가공 /ex)2022.7월
#
# sheet_list 를 이용해서 월별 시트 생성 / xl_style 적용
# school_list를 이용해서 학교 구분해서 엑셀에 입력
#######################################Input Data Control Part#######################################

#이 값이 바뀌면 비교 키값인 school_list와 sheet_list의 값 검증이 필요할 수 있음
wln = 1 #Write_List_Name - 이름의 colume 값 위치
wld = 2 #Write_List_Date - 날짜의 colume 값 위치

#문서 작성시 빈칸으로 떨어져야할 cell 값
Doc_style_col = 1 #빈칸으로 들어갈 colume 값

#열 너비 값 [C: 점검일자, D: 신청자, E: 근무실, F: 증상]
xl_style_column = ['C','D','E','F']
xl_style_column_value = ['16','12','16','36'] # [default : 8.38] / C: 15.88 / D: 11.13 / E: 15.88 / F: 36.13

#xl_Style Default Setting Part에서 그려질 표의 행 수
xl_style_table_num = 80

#표 스타일 지정
box = Border(left   = Side(border_style='thin', color='000000'),
             right  = Side(border_style='thin', color='000000'),
             top    = Side(border_style='thin', color='000000'),
             bottom = Side(border_style='thin', color='000000'))

#####################################################################################################

#Write_List 엑셀 읽어오기
print("Write_List 엑셀 읽어오는중...")
wb = openpyxl.load_workbook('./Write_List.xlsx')
#Write_List Sheet 지정
ws = wb['Sheet1']

#반복문 - Write_List 값 List에 추가
get_cells = ws
row_list = []
for row in get_cells:
    cell_list = []
    for i in range(0,Doc_style_col):
        cell_list.insert(0, None)
    for cell in row:
        cell_list.append(cell.value)
    row_list.append(cell_list)

#school_list 생성 - Write_List의 학교명 확인을 위한 변수
print("학교명 확인중...")
school_list = []
for i in range(1, ws.max_row):
    school_list.append(row_list[i][wln]) 

#school_list 중복제거 / rwr = repeated word removal(반복 단어 제거)
rwr_array = []
rwr_array = set(school_list)
school_list = list(rwr_array)
school_list.sort()

#key 변수 선언 및 school_list의 엑셀 호출
print("학교 엑셀 읽어오는중...")
error_school_list = '초기화'
try:                                        #error 발생시 except: 구문을 실행
    for i in range(0,len(school_list)):
        error_school_list = school_list[i]
        globals()['key_{}'.format(i+1)] = openpyxl.load_workbook('./'+school_list[i]+'.xlsx')
except:
    messagebox.showerror("error message", "불러올 " + error_school_list + " excel이 없습니다.") #widows에서 messagebox로 error 창 출력


#sheet_list 생성 - 월별 Sheet 판별을 위한 변수
sheet_list = []
for i in range(1, ws.max_row):
    sheet_list.append(row_list[i][wld])

#sheet_list 문자열 정리 ex) 20220804 -> 202208
for i in range(0,len(sheet_list)):
    slicing = str(sheet_list[i])
    slicing = slicing[:6]
    sheet_list[i] = slicing

#sheet_list 중복제거 및 오름차순 정렬 / rwr_array 위에서 정의함 
rwr_array = set(sheet_list)
sheet_list = list(rwr_array)
sheet_list.sort()

#sheet_list_name 변수 선언 및 가공 ex) 20220804 -> 2022.08월
sheet_list_name = []
for i in range(0,len(sheet_list)):
    temp = sheet_list[i]
    sheet_list_name.append(temp)
for i in range(0,len(sheet_list_name)):
    slicing_first = sheet_list_name[i]
    slicing_first = slicing_first[:4]
    slicing_second = sheet_list_name[i]
    slicing_second = slicing_second[4:]
    if slicing_second[0] == '0':
        slicing_second = slicing_second[1:]
    sheet_list_name[i] = slicing_first + '.' + slicing_second + '월'

###시트 생성 Part###
print("시트 생성중...")
#key_sheet 변수 선언 / sheet_list만큼의 sheet 생성 / xl_style 설정
for i in range(0,len(school_list)):
    #sheet_temp1는 key값 받아서 sheet지정하기위한 변수 / key_sheet_[i+1] = temp[sheet_list]
    num = 0                                                                #sheet_list[0] <- 기존 작성중인 sheet가 있을 경우를 위해서 설정하는 값
    sheet_temp1 = globals()['key_{}'.format(i+1)]

    #tempnames는  sheet_list가 있는지 확인하고 생성하기 위한 변수
    tempnames = []
    tempnames = sheet_temp1.sheetnames
    
    #기존에 작성중인 sheet가 있는 경우
    if sheet_list_name[0] in tempnames: #sheet_list_name[0] 값이 sheet에 있는지 판별하는 조건문
        num = 1                         #기존 작성중인 sheet가 존재할 시 sheet_list_name[0]다음 것부터 sheet 생성하게 하기 위한 변수
        
        ###기존 sheet의 None값(표) 삭제
        globals()['key_sheet_{}'.format(i+1)] = globals()['key_{}'.format(i+1)][sheet_list_name[0]] #sheet를 sheet_list_name[0]으로 지정
        
        #None값 위치 찾기
        none_row = 8 # none_row 변수 선언
        for j in range(8,xl_style_table_num+1):
            if globals()['key_sheet_{}'.format(i+1)]['C'+str(none_row)].value == None: #None일 경우
                globals()['key_sheet_{}'.format(i+1)].delete_rows(none_row)
            else:
                none_row+=1

    for j in range(num, len(sheet_list_name)):
        sheet_temp1.create_sheet(sheet_list_name[j])                       #school_list값 별로 sheet에 sheet_list_name의 값으로 sheet를 생성
        globals()['key_{}'.format(i+1)] = sheet_temp1                      #생성한 sheet값은 임시저장한 sheet_temp1에 저장된 것이니깐 변수 key_i에 저장
        globals()['key_sheet_{}'.format(i+1)] = globals()['key_{}'.format(i+1)][sheet_list_name[j]] #생성한 key_sheet_i 변수 생성 및 불러온 엑셀의 작업 sheet 지정

        ###xl_Style Default Setting Part###
        sheet_temp2 = sheet_list[j]
        slicing_mm = sheet_temp2[-2:]
        if slicing_mm[0] == '0':
            slicing_mm = slicing_mm[1:]
        xl_style_temp = globals()['key_sheet_{}'.format(i+1)]

        for l in range(0, len(xl_style_column)):
            xl_style_temp.column_dimensions[xl_style_column[l]].width = int(xl_style_column_value[l])

        #title
        xl_style_temp.merge_cells("C2:F3")
        xl_style_temp['C2']            = "정보화기기 유지보수 정기점검 기록부 ("+ slicing_mm +"월)"
        xl_style_temp['C2'].font       = Font(size=20, bold = True)
        xl_style_temp['C2'].alignment  = Alignment(horizontal='center', vertical='center') 
        
        #schoole_name
        xl_style_temp['B5']            = "학교명:"
        xl_style_temp['B5'].font       = Font(size=12, bold = True)
        
        modify_school_list             = school_list[i]
        if '초' in modify_school_list:
            modify_school_list             = modify_school_list[:modify_school_list.index('초')+1]
            modify_school_list = modify_school_list +'등학교'
        elif '중' in modify_school_list:
            modify_school_list             = modify_school_list[:modify_school_list.index('중')+1]
            modify_school_list = modify_school_list +'학교'
        xl_style_temp['C5']            = modify_school_list
        xl_style_temp['C5'].font       = Font(size=12, bold = True)

        #menu
        menu = Font(size=16, bold = True) 
        menu_num = ['B7','C7','D7', 'E7', 'F7', 'G7']
        menu_value = ['연번', '점검일자', '신청자', '근무실', '증상', '비고']
        for l in range(0,len(menu_num)):
            xl_style_temp[menu_num[l]]           = menu_value[l]
            xl_style_temp[menu_num[l]].font      = menu
            xl_style_temp[menu_num[l]].border    = box
            xl_style_temp[menu_num[l]].alignment = Alignment(horizontal='center', vertical='center')

###값 입력 Part###
print("Write_List 값 입력중...")        
#조건문 - Write_List값을 저장한 row_list의 값을 날짜와 학교를 비교해서 각 엑셀로 분배
for i in range(1, ws.max_row):              #Write_List의 행 수만큼 반복
    slicing = str(row_list[i][wld])
    slicing = slicing[:6]
    for j in range(0, len(sheet_list)):     #일치하는 월 Sheet를 찾기위해서 sheet_list만큼 반복
        if slicing == sheet_list[j]:        #입력 날짜와 sheet의 월이 일치하는지 검증
            for l in range(0, len(school_list)):        #일치하는 school을 찾기위해 school_list 만큼 반복
                if row_list[i][wln] == school_list[l]:  #입력 school과 school_list의 값이 일치하는지 검즘
                    globals()['key_sheet_{}'.format(j+1)] = globals()['key_{}'.format(l+1)][sheet_list_name[j]] #school과 sheet를 지정
                    modify_list = row_list[i]
                    del modify_list[1]      
                    modify_list.insert(1,globals()['key_sheet_{}'.format(j+1)].max_row-6)
                    modify_list[2] = str(modify_list[2])[:4] + "-" + str(modify_list[2])[4:6] +  "-" + str(modify_list[2])[-2:]
                    globals()['key_sheet_{}'.format(j+1)].append(modify_list)

                    ###Add Value and xl_Style Setting Part###
                    value_num = ['B'+str(globals()['key_sheet_{}'.format(j+1)].max_row),
                                 'C'+str(globals()['key_sheet_{}'.format(j+1)].max_row),
                                 'D'+str(globals()['key_sheet_{}'.format(j+1)].max_row),
                                 'E'+str(globals()['key_sheet_{}'.format(j+1)].max_row),
                                 'F'+str(globals()['key_sheet_{}'.format(j+1)].max_row),
                                 'G'+str(globals()['key_sheet_{}'.format(j+1)].max_row)]

                    for k in range(0,len(value_num)):
                        globals()['key_sheet_{}'.format(j+1)][value_num[k]].border    = box
                        globals()['key_sheet_{}'.format(j+1)][value_num[k]].alignment = Alignment(horizontal='center', vertical='center')
                        globals()['key_sheet_{}'.format(j+1)].row_dimensions[globals()['key_sheet_{}'.format(j+1)].max_row].height = 22

###빈 표 입력 Part###
for i in range(0,len(school_list)):
    for j in range(0,len(sheet_list_name)):
        globals()['key_sheet_{}'.format(j+1)] = globals()['key_{}'.format(i+1)][sheet_list_name[j]]
        for l in range(globals()['key_sheet_{}'.format(j+1)].max_row+1, xl_style_table_num+1):
            table_num = ['B'+str(l),
                         'C'+str(l),
                         'D'+str(l),
                         'E'+str(l),
                         'F'+str(l),
                         'G'+str(l)]
            for k in range(0, len(table_num)):
                if k == 0:
                    globals()['key_sheet_{}'.format(j+1)][table_num[k]].value = l-7
                globals()['key_sheet_{}'.format(j+1)][table_num[k]].border = box
                globals()['key_sheet_{}'.format(j+1)][table_num[k]].alignment = Alignment(horizontal='center', vertical='center')
                globals()['key_sheet_{}'.format(j+1)].row_dimensions[l].height = 22

#파일 저장
print("파일 저장중...")
for i in range(0, len(school_list)):
    globals()['key_{}'.format(i+1)].save('./'+school_list[i] +'.xlsx')